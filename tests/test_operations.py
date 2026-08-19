import json
from pathlib import Path

import numpy as np
import pytest

from fraud_strategy.operations import (
    OPERATING_CAPACITY,
    assign_bands,
    band_edges,
    period_date,
    period_date_key,
)
from fraud_strategy.suspects import RING_MAXIMUM, RING_MINIMUM, day_of


def load(name: str) -> dict:
    return json.loads(Path(name).read_text(encoding="utf-8"))


def test_risk_bands_are_fixed_on_the_calibration_period_not_the_batch() -> None:
    """Guard the M6 lesson, applied to bands.

    The champion's score mapping was defective because it standardised against whatever
    batch it received. A band that re-derives its edges per batch has the same defect:
    "critical" would mean a different thing every period, which is useless to an
    investigator. Edges are fitted once and then applied to unseen, shifted data.
    """
    calibration = np.linspace(0.0, 1.0, 10_001)
    edges = band_edges(calibration)

    shifted = np.clip(calibration + 0.30, 0, 1)
    bands = assign_bands(shifted, edges)
    reused = assign_bands(calibration, edges)

    # Same edges, different populations, so the shifted period must band more severely.
    assert (bands == "critical").sum() > (reused == "critical").sum()
    # And the edges themselves are untouched by the second population.
    assert band_edges(calibration) == edges


def test_bands_are_ordered_and_cover_every_score() -> None:
    edges = band_edges(np.linspace(0, 1, 1_001))
    assert edges["critical"] > edges["high"] > edges["medium"]
    bands = assign_bands(np.array([0.0, edges["medium"], edges["high"], edges["critical"], 1.0]), edges)
    assert list(bands) == ["low", "medium", "high", "critical", "critical"]


def test_periods_map_to_sortable_keys_and_never_claim_a_calendar_month() -> None:
    keys = [period_date_key(period) for period in range(8)]
    assert keys == sorted(keys)
    assert len(set(keys)) == 8
    assert period_date(0).day == 1
    # Twelve periods must roll the year rather than produce month 13.
    assert period_date(12).year == period_date(0).year + 1


def test_suspect_days_partition_every_application_exactly_once() -> None:
    applications, days = 50_000, 20
    counts: dict[int, int] = {}
    for index in range(applications):
        day = day_of(index, applications, days)
        assert 0 <= day < days
        counts[day] = counts.get(day, 0) + 1
    assert sum(counts.values()) == applications
    assert day_of(0, applications, days) == 0
    assert day_of(applications - 1, applications, days) == days - 1


def test_monthly_kpi_pack_reads_the_database_and_keeps_the_refusal() -> None:
    pack = load("evaluation/monthly_kpi.json")
    assert "PostgreSQL" in pack["source"]
    assert pack["review_capacity"] == OPERATING_CAPACITY
    assert len(pack["periods"]) == 8
    # Both models present in every period, and neither is promoted.
    assert len(pack["monthly"]) == 16
    assert {run["recommendation"] for run in pack["scenario_runs"]} == {"no robust recommendation"}
    assert {run["approval_state"] for run in pack["scenario_runs"]} == {"rejected", "retained_baseline"}


def test_monthly_kpi_arithmetic_is_internally_consistent() -> None:
    for row in load("evaluation/monthly_kpi.json")["monthly"]:
        assert row["fraud_caught"] + row["fraud_missed"] == row["fraud_attempts"]
        assert row["fraud_caught"] <= row["queue_size"]
        assert 0 <= row["catch_rate"] <= 1
        assert row["capacity_overshoot"] == row["queue_size"] - row["capacity_headcount"]
        assert abs(row["catch_rate"] - row["fraud_caught"] / row["fraud_attempts"]) < 1e-6


def test_a_tied_score_overshoots_review_capacity_and_the_pack_says_so() -> None:
    """Guard the M9 finding.

    The incumbent proxy is a low-cardinality integer score, so the block of applications
    sitting exactly on the cut cannot be split and the queue runs over the staffed
    headcount. A desk paying for those reviews needs the number, so it must not be
    silently rounded away.
    """
    rows = load("evaluation/monthly_kpi.json")["monthly"]
    incumbent = [row for row in rows if row["model_version"].startswith("incumbent")]
    challenger = [row for row in rows if not row["model_version"].startswith("incumbent")]
    assert incumbent and challenger
    assert all(row["capacity_overshoot"] > 0 for row in incumbent)
    assert max(row["capacity_overshoot"] for row in challenger) <= 1
    assert min(row["capacity_overshoot"] for row in incumbent) > 10


def test_vendor_comparison_holds_review_capacity_equal() -> None:
    pack = load("evaluation/monthly_kpi.json")
    assert len(pack["vendor_performance"]) == 8
    for row in pack["vendor_performance"]:
        assert row["review_capacity"] == OPERATING_CAPACITY
        assert row["challenger_catch_rate"] > row["incumbent_catch_rate"]
        assert row["additional_fraud_caught"] > 0


def test_daily_suspect_report_never_claims_a_link_in_the_source_data() -> None:
    report = load("evaluation/daily_suspect_report.json")
    assert report["evidence_source"] == "synthetic_link_fixture"
    assert "no BAF application is linked" in report["boundary"]
    assert "no BAF row is linked or claimed" in report["validation"]["scope"]
    serialized = json.dumps(report)
    assert "baf_base" not in serialized


def test_daily_suspect_report_produces_an_operational_cadence() -> None:
    report = load("evaluation/daily_suspect_report.json")
    assert report["totals"]["days"] == len(report["days"])
    assert sum(day["applications_received"] for day in report["days"]) == report["totals"]["applications"]
    assert report["totals"]["duplicate_identity_flags"] > 0
    assert report["totals"]["suspected_ring_flags"] > 0
    # Ring groups stay inside the window the linking evaluation uses.
    for flag in report["latest_day_flags"]:
        if flag["type"] == "suspected_ring":
            assert RING_MINIMUM <= flag["cluster_size"] <= RING_MAXIMUM


def test_daily_suspect_report_validates_against_withheld_fixture_truth() -> None:
    validation = load("evaluation/daily_suspect_report.json")["validation"]
    assert validation["duplicate_pair_precision"] == 1.0
    assert validation["duplicate_pair_recall"] == 1.0
    assert validation["ring_recall"] == 1.0
    assert validation["false_ring_groups"] == 0


def _fairness(tpr_gap: float, fpr_gap: float, ratio: float) -> dict:
    return {
        "segments": [
            {
                "segment": "housing_status",
                "max_min_tpr_gap": tpr_gap,
                "max_min_fpr_gap": fpr_gap,
                "review_rate_ratio": ratio,
            }
        ],
        "governance_review": True,
        "reasons": [
            "housing_status max_min_tpr_gap exceeds 0.10",
            "housing_status review-rate ratio is outside 0.80-1.25",
        ],
    }


def _acceptance(reasons: list[str]) -> dict:
    return {
        "accepted_reasons": reasons,
        "granted_against": {
            "housing_status": {
                "max_min_tpr_gap": 0.3771,
                "max_min_fpr_gap": 0.1556,
                "review_rate_ratio": 0.0732,
            }
        },
        "tolerance": {"gap": 0.05, "ratio": 0.05},
    }


def test_a_recorded_acceptance_resolves_the_gate_it_was_written_for() -> None:
    """Gate 5 reads 'resolved or explicitly accepted by governance'.

    Only the first branch was ever computed, so a recorded human acceptance could not
    satisfy the gate written to receive it. This covers the second branch.
    """
    from fraud_strategy.modeling import apply_governance_acceptance

    base = _fairness(0.3771, 0.1556, 0.0732)
    result = apply_governance_acceptance(base, _acceptance(base["reasons"]))
    assert result["governance_review"] is False
    assert result["outstanding_reasons"] == []
    assert len(result["accepted_reasons"]) == 2
    # The warnings stay visible. Accepting is not hiding.
    assert result["reasons"] == base["reasons"]


def test_an_acceptance_does_not_cover_a_reason_it_never_named() -> None:
    from fraud_strategy.modeling import apply_governance_acceptance

    base = _fairness(0.3771, 0.1556, 0.0732)
    partial = _acceptance(["housing_status max_min_tpr_gap exceeds 0.10"])
    result = apply_governance_acceptance(base, partial)
    assert result["governance_review"] is True
    assert result["outstanding_reasons"] == ["housing_status review-rate ratio is outside 0.80-1.25"]


def test_an_acceptance_lapses_when_the_measured_condition_worsens() -> None:
    """An acceptance that cannot lapse is a permanent exemption wearing a governance label.

    It is a judgment about a measured condition, so it has to expire when that condition
    moves materially against the people it affects.
    """
    from fraud_strategy.modeling import apply_governance_acceptance

    base = _fairness(0.3771, 0.1556, 0.0732)
    accepted = _acceptance(base["reasons"])

    widened = apply_governance_acceptance(_fairness(0.4400, 0.1556, 0.0732), accepted)
    assert widened["governance_review"] is True
    assert widened["lapsed_reasons"]

    fell = apply_governance_acceptance(_fairness(0.3771, 0.1556, 0.0100), accepted)
    assert fell["governance_review"] is True

    # Inside tolerance, and in the improving direction, the acceptance still holds.
    steady = apply_governance_acceptance(_fairness(0.3800, 0.1500, 0.0900), accepted)
    assert steady["governance_review"] is False


def test_no_acceptance_file_leaves_the_gate_exactly_as_it_was() -> None:
    from fraud_strategy.modeling import apply_governance_acceptance

    base = _fairness(0.3771, 0.1556, 0.0732)
    result = apply_governance_acceptance(base, None)
    assert result["governance_review"] is True
    assert result["outstanding_reasons"] == base["reasons"]
    assert result["accepted_reasons"] == []


def test_the_recorded_acceptance_names_the_segments_the_report_flags() -> None:
    acceptance = load("evaluation/governance_acceptance.json")
    flagged = load("evaluation/model_evaluation.json")["fairness"]["reasons"]
    assert acceptance["decision"] == "accepted"
    assert acceptance["accountable"]
    assert acceptance["date"] == "2026-08-10"
    assert set(acceptance["accepted_reasons"]) == set(flagged)
    assert "retained" in acceptance["feature_decision"]
    # It resolves gate 5 and says so; it must not claim to resolve the others.
    assert any("gate 5 only" in limit for limit in acceptance["limits"])


def test_the_originations_strategy_states_a_refusal_and_matches_the_evidence() -> None:
    """Guard the document against evidence drift.

    A strategy document that quotes stale figures is worse than no document, because it
    reads as authoritative. These are the figures the recommendation rests on.
    """
    text = Path("docs/originations-strategy.md").read_text(encoding="utf-8")
    model = load("evaluation/model_evaluation.json")
    capacity = model["month_7_capacity"]

    for name, expected in (
        ("incumbent_proxy", "20.59%"),
        ("catboost_hybrid", "53.64%"),
        ("regularized_logistic", "49.37%"),
    ):
        measured = capacity[name]["0.05"]["catch_rate"]
        assert f"{measured:.2%}" == expected
        assert expected in text

    caught = capacity["catboost_hybrid"]["0.05"]["fraud_caught"]
    incumbent_caught = capacity["incumbent_proxy"]["0.05"]["fraud_caught"]
    assert str(caught - incumbent_caught) in text, "the headline gain must match the evidence"

    # It must not read as an approval, and both failing gates must be named.
    assert "Do not promote" in text
    for gate, passed in model["promotion_gates"].items():
        if not passed:
            assert gate.split("_")[0] in text.lower() or "stability" in text.lower()
    assert model["strategy_frontier"]["recommendation"] == "no robust recommendation"


def test_the_strategy_document_does_not_quote_a_money_figure_in_its_recommendation() -> None:
    """The refusal rests on calibration and stability, not on economics.

    A dollar amount in the recommendation would misrepresent why the answer is no, and
    would invite the reader to argue the refusal away on price. The money belongs in the
    option comparison, where it is stated as a range.
    """
    text = Path("docs/originations-strategy.md").read_text(encoding="utf-8")
    recommendation = text.split("## 5. Recommendation")[1].split("## 6.")[0]
    assert "$" not in recommendation


def test_the_promotion_gate_still_uses_the_approved_grid_unchanged() -> None:
    """Adding sensitivity dimensions must not widen the grid a gate is measured against.

    "Positive for at least 80% of the approved assumption grid" means the 60-point grid the
    M2 contract approved. Sweeping two more factors through it would change what the gate
    asserts, which needs a new approved contract. Same standard M6 applied to the intercept.
    """
    from fraud_strategy.strategy import compare_utility_grid, economic_sensitivity_surface

    challenger = {"fraud_caught": 766, "false_positive_count": 4076, "queue_size": 4842}
    incumbent = {"fraud_caught": 294, "false_positive_count": 4548, "queue_size": 4842}

    gate = compare_utility_grid(challenger, incumbent)
    assert gate["grid_points"] == 60
    assert gate["positive_share"] == 1.0
    assert gate["incremental_utility_min"] == 2_360_000.0

    surface = economic_sensitivity_surface(challenger, incumbent)
    assert surface["status"] == "sensitivity evidence, not a promotion gate"
    assert all(cell["grid_points"] == 60 for cell in surface["cells"])
    # The unadjusted corner of the surface must reproduce the gate exactly.
    unadjusted = next(
        cell
        for cell in surface["cells"]
        if cell["review_effectiveness"] == 1.0 and cell["loss_given_fraud"] == 1.0
    )
    for key in ("positive_share", "incremental_utility_min", "incremental_utility_max"):
        assert unadjusted[key] == gate[key]


def test_the_economic_sign_holds_across_the_whole_sensitivity_surface() -> None:
    """The decision is robust even though the headline is not.

    At equal capacity the challenger catches more fraud AND holds up fewer good customers,
    and review cost cancels, so scaling only the benefit term cannot reverse it. If this
    ever fails, the recommendation in docs/originations-strategy.md has to be rewritten.
    """
    from fraud_strategy.strategy import economic_sensitivity_surface

    surface = economic_sensitivity_surface(
        {"fraud_caught": 766, "false_positive_count": 4076, "queue_size": 4842},
        {"fraud_caught": 294, "false_positive_count": 4548, "queue_size": 4842},
    )
    assert surface["sign_holds_everywhere"] is True
    assert surface["reported_band"]["min"] < surface["reported_band"]["max"]
    assert surface["inflation_factor_at_worst_case"] > 1.0


def test_neither_recovery_factor_is_presented_as_an_observed_figure() -> None:
    from fraud_strategy.strategy import economic_sensitivity_surface

    surface = economic_sensitivity_surface(
        {"fraud_caught": 766, "false_positive_count": 4076, "queue_size": 4842},
        {"fraud_caught": 294, "false_positive_count": 4548, "queue_size": 4842},
    )
    assert "citation bar" in surface["sourcing"]
    assert "declared sensitivity inputs" in surface["sourcing"]
    assert max(surface["review_effectiveness_range"]) == 1.0
    assert max(surface["loss_given_fraud_range"]) == 1.0


def test_the_walk_forward_final_fold_reproduces_the_recorded_protocol() -> None:
    """The harness has to validate itself before its other numbers mean anything.

    Fold 7 trains on 0-5, calibrates on 6 and tests on 7, which is the recorded evaluation
    protocol exactly. If it disagrees with the recorded month-7 result, the backtest is
    wrong, not the record.
    """
    backtest = load("evaluation/walk_forward_backtest.json")
    recorded = load("evaluation/model_evaluation.json")
    assert backtest["summary"]["final_fold_reproduces_recorded_protocol"] is True

    final = backtest["folds"][-1]
    assert final["test_period"] == 7
    assert final["train_periods"] == [0, 1, 2, 3, 4, 5]
    assert final["calibration_period"] == 6
    expected = recorded["month_7_capacity"]["catboost_hybrid"]["0.05"]["catch_rate"]
    assert abs(final["catboost_hybrid"]["catch_rate"]["0.05"] - expected) < 1e-9


def test_no_walk_forward_fold_trains_on_the_untouched_period() -> None:
    backtest = load("evaluation/walk_forward_backtest.json")
    for fold in backtest["folds"]:
        assert 7 not in fold["train_periods"]
        assert fold["calibration_period"] != 7
        assert fold["calibration_period"] == fold["test_period"] - 1
        assert max(fold["train_periods"]) == fold["test_period"] - 2
    assert backtest["protocol"]["no_fold_trains_on_the_untouched_period"] is True


def test_the_recorded_interval_understates_next_period_uncertainty() -> None:
    """The corrected version of an earlier claim.

    Pre-work estimated between-period variation at 1.74 times the within-period interval,
    using catch rates from months the model had trained on. Measured properly out of
    sample the two are about equal. That does not rescue the recorded interval: a decision
    about next period faces both sources at once, which is what the prediction interval
    combines, and it is materially wider than what the record currently reports.
    """
    summary = load("evaluation/walk_forward_backtest.json")["summary"]
    assert len(summary["periods_tested"]) == 5

    recorded_width = summary["recorded_single_period"]["within_period_interval_width"]
    predict_low, predict_high = summary["next_period_prediction_interval_95"]
    assert (predict_high - predict_low) > 2 * recorded_width

    # And the recorded single period sits above the out-of-sample mean, so the headline
    # was taken from a better-than-average period.
    assert summary["recorded_single_period"]["catch_rate"] > summary["mean_catch_rate"]


def test_the_prediction_interval_is_wider_than_the_interval_for_the_mean() -> None:
    """They answer different questions and the wider one is the operational answer."""
    summary = load("evaluation/walk_forward_backtest.json")["summary"]
    mean_low, mean_high = summary["mean_catch_rate_interval_95"]
    predict_low, predict_high = summary["next_period_prediction_interval_95"]
    assert predict_low < mean_low
    assert predict_high > mean_high
    assert "prediction interval is the one a next-period forecast needs" in summary["interval_note"]


def test_every_walk_forward_fold_reports_a_usable_catch_rate() -> None:
    backtest = load("evaluation/walk_forward_backtest.json")
    for fold in backtest["folds"]:
        assert fold["positives"] > 200, "a period with too few positives cannot support a rate"
        for model in ("catboost_hybrid", "incumbent_proxy"):
            for capacity, value in fold[model]["catch_rate"].items():
                assert 0.0 < value < 1.0, f"{model} period {fold['test_period']} at {capacity}"
        # The challenger must beat the incumbent in every period, or the strategy
        # document's dominance claim does not survive out of sample.
        assert fold["catboost_hybrid"]["catch_rate"]["0.05"] > fold["incumbent_proxy"]["catch_rate"]["0.05"]


def test_two_calibration_metrics_on_the_same_predictions_disagree() -> None:
    """The evidence for M6's parameterisation hypothesis, now measurable.

    Expected calibration error reads across the score distribution the desk actually sees.
    The intercept reads the recalibration line at probability 0.5, far above where any
    application sits. If these two ever start agreeing, the argument in section 6 of the
    strategy document has to be withdrawn.
    """
    folds = load("evaluation/walk_forward_backtest.json")["folds"]
    observations = [fold[model] for fold in folds for model in ("catboost_hybrid", "incumbent_proxy")]
    assert len(observations) == 10
    assert all(item["ece"] <= 0.02 for item in observations), "ECE passes everywhere"
    assert all(abs(item["calibration_intercept"]) > 0.10 for item in observations), (
        "intercept fails everywhere"
    )
    # The gap is not marginal in either direction.
    assert max(item["ece"] for item in observations) < 0.005
    assert min(abs(item["calibration_intercept"]) for item in observations) > 0.20


def test_recalibrating_on_the_prior_period_does_not_fix_the_intercept() -> None:
    """Guard the corrected recommendation.

    Every fold calibrates on the period immediately before its test, which is exactly the
    remedy an earlier version of the strategy document proposed. It does not work, and the
    document must not drift back to proposing it.
    """
    backtest = load("evaluation/walk_forward_backtest.json")
    assert backtest["summary"]["calibration_intercept_fails_in_every_period"] is True
    for fold in backtest["folds"]:
        assert fold["calibration_period"] == fold["test_period"] - 1

    # Normalised, because the document is hard-wrapped and a prose assertion should not
    # fail on where a line happens to break.
    strategy = " ".join(Path("docs/originations-strategy.md").read_text(encoding="utf-8").split())
    assert "recalibrating on the preceding period does not bring the intercept inside the limit" in strategy
    assert "standing limitation" in strategy


def test_the_recalibration_trigger_has_no_input_at_period_close() -> None:
    """Guard the M8 finding against the M6 control.

    The control reads the just-closed period's observed prior. That cohort has had no
    payment fall due, so its maturity is zero and there is no number to read. If this
    ever passes, someone has quietly assumed labels arrive faster than they do.
    """
    latency = load("evaluation/label_latency.json")
    at_close = next(item for item in latency["censored_trigger"] if item["reporting_lag_periods"] == 0)
    assert at_close["cohort_maturity"] == 0.0
    assert at_close["decisions_available"] == 0


def test_censoring_bias_exceeds_the_trigger_threshold_at_usable_lags() -> None:
    latency = load("evaluation/label_latency.json")
    threshold = latency["assumptions"]["trigger_threshold_logit"]
    one_period = next(item for item in latency["censored_trigger"] if item["reporting_lag_periods"] == 1)
    assert abs(one_period["differential_bias_logit"]) > 5 * threshold
    # Bias is negative at every lag: a newer cohort is always less mature than the one it
    # is compared against, so the observed move always understates the real one.
    for item in latency["censored_trigger"]:
        if item["differential_bias_logit"] is not None:
            assert item["differential_bias_logit"] < 0
    assert latency["minimum_lag_where_bias_is_below_the_threshold"] >= 3


def test_correcting_by_the_stated_maturity_restores_the_trigger() -> None:
    latency = load("evaluation/label_latency.json")
    for item in latency["corrected_trigger"]:
        assert item["decisions_agreeing_with_true_priors"] == item["decisions_available"]
        assert item["decisions_available"] > 0
    # And it is labelled as a stated model rather than a measurement, because this dataset
    # has no label timestamps to validate the curve against.
    assert "stated model, not measured" in latency["status"]


def test_hit_rate_is_a_leading_indicator_and_catch_rate_is_not() -> None:
    """They share a numerator; only the denominator is slow.

    A review confirms fraud within days. Catch rate divides that same numerator by all
    fraud in the period, including what slipped past and defaults months later.
    """
    signals = {
        item["signal"]: item
        for item in load("evaluation/label_latency.json")["observable_inside_the_latency_window"]
    }
    hit_rate = next(item for name, item in signals.items() if "hit rate" in name.lower())
    catch_rate = next(item for name, item in signals.items() if name == "Catch rate")
    assert hit_rate["latency"] == "days"
    assert "30 to 90 days" in catch_rate["latency"]
    assert "denominator" in catch_rate["why"]


def test_the_matcher_is_blind_to_a_pair_that_varies_email_and_phone() -> None:
    """The blind spot, measured rather than asserted.

    Corrupting only the two signals the acceptance rule requires destroys recall.
    Corrupting the other two at identical rates does nothing at all. The asymmetry is the
    proof that the cause is the rule and not corruption in general.
    """
    adversarial = load("evaluation/linking_adversarial.json")
    findings = adversarial["findings"]
    assert findings["targeted_corruption_collapses_recall"] is True
    assert findings["control_corruption_holds_recall"] is True
    assert findings["recall_at_full_targeted_corruption"] < 0.10
    assert findings["recall_at_matching_control_corruption"] == 1.0

    for row in adversarial["control_corruption"]:
        assert row["recall"] == 1.0, "corrupting device and address must not move recall"
    worst = min(adversarial["targeted_corruption"], key=lambda row: row["recall"])
    assert worst["corruption_rate"] == 1.0


def test_the_blind_spot_is_recorded_with_the_reason_the_rule_exists() -> None:
    """A trade recorded without its rationale reads as a defect. This one is a choice."""
    rule = load("evaluation/linking_adversarial.json")["acceptance_rule"]
    assert "cannot be accepted at any score" in rule["blind_spot"]
    assert "false merges" in rule["why_it_exists"]
    assert rule["signal_weights"]["email"] + rule["signal_weights"]["phone"] > 0.5


def test_the_base_rate_test_reports_that_it_cannot_measure_precision() -> None:
    """A clean result from a test that cannot fail is not evidence.

    Precision holds at every base rate because the fixture contains no near-misses. That
    is a statement about the fixture, and the evidence file has to say so rather than
    presenting an uninformative pass as a strength.
    """
    findings = load("evaluation/linking_adversarial.json")["findings"]
    assert findings["base_rate_test_is_informative"] is False
    assert "statement about the fixture" in findings["base_rate_interpretation"]
    assert "hard negatives" in findings["base_rate_interpretation"]


def test_the_linking_analysis_makes_no_claim_about_the_source_data() -> None:
    adversarial = load("evaluation/linking_adversarial.json")
    assert adversarial["evidence_source"] == "synthetic_link_fixture"
    assert "no BAF application is linked" in adversarial["boundary"]
    assert "baf_base" not in json.dumps(adversarial)


def test_reject_inference_is_recorded_as_a_position_with_a_precondition() -> None:
    """The counterfactual cannot be recovered after the fact, so the design has to precede
    the first binding decision. If this document ever stops saying that, the mitigation has
    become retrofittable in someone's mind, and it is not."""
    text = " ".join(Path("docs/reject-inference.md").read_text(encoding="utf-8").split())
    assert "The product is non-binding" in text
    assert "random-approval control group" in text.lower() or "A random-approval control group" in text
    assert "before go-live" in text or "before the first binding decision" in text
    assert "cannot be recovered once it is lost" in text
    # And it must own the inherited limitation rather than claiming unconditional results.
    assert "undocumented" in text and "selection process" in text


def test_a_refit_changes_the_file_hash_while_behaviour_is_identical() -> None:
    """The M9 open finding, verified rather than restated, and the reason for M12's fix.

    Two fits over identical data with the same seed produce different bytes, because
    CatBoost writes fit-time metadata into the model, while their predictions are
    bit-identical. A rollback decision made on the artifact hash would therefore conclude
    the model had changed when nothing about it had. The fingerprint hashes behaviour and
    stays put.

    Re-saving an already-fitted model is byte-stable, which was checked while writing this
    and is worth knowing: the instability is specific to refitting, not to serialisation.
    """
    import tempfile

    from catboost import CatBoostClassifier

    from fraud_strategy.io import sha256_file
    from fraud_strategy.modeling import behaviour_fingerprint

    generator = np.random.default_rng(7)
    features = generator.normal(size=(2_000, 6))
    labels = (generator.random(2_000) < 0.05).astype(int)

    def fit_once(path: Path) -> tuple[str, np.ndarray]:
        model = CatBoostClassifier(
            iterations=30, depth=3, random_seed=20260805, allow_writing_files=False, verbose=False
        )
        model.fit(features, labels)
        model.save_model(path.as_posix())
        return sha256_file(path), model.predict_proba(features)[:, 1]

    with tempfile.TemporaryDirectory() as directory:
        first_hash, first_scores = fit_once(Path(directory) / "first.cbm")
        second_hash, second_scores = fit_once(Path(directory) / "second.cbm")

        # Re-saving the same fitted model does not move the hash. Only refitting does.
        resave = Path(directory) / "resave.cbm"
        model = CatBoostClassifier(
            iterations=30, depth=3, random_seed=20260805, allow_writing_files=False, verbose=False
        )
        model.fit(features, labels)
        model.save_model(resave.as_posix())
        repeat = Path(directory) / "repeat.cbm"
        model.save_model(repeat.as_posix())
        assert sha256_file(resave) == sha256_file(repeat)

    assert first_hash != second_hash, "the file hash moves on refit; that is the problem"
    assert np.array_equal(first_scores, second_scores), "behaviour does not move"
    assert behaviour_fingerprint(first_scores) == behaviour_fingerprint(second_scores)


def test_the_manifests_carry_a_fingerprint_that_separates_the_two_models() -> None:
    candidate_path = Path("artifacts/models/candidate_model_manifest.json")
    champion_path = Path("artifacts/models/model_manifest.json")
    manifests = (candidate_path, champion_path)
    if not all(path.is_file() for path in manifests):
        pytest.skip("governed private model manifests are not distributed with the public repository")

    candidate = load(candidate_path.as_posix())["behaviour_fingerprint"]
    champion = load(champion_path.as_posix())["behaviour_fingerprint"]
    for record in (candidate, champion):
        assert len(record["digest"]) == 64
        assert record["reference_rows"] == 5_000
        assert record["reference_period"] == 6
        assert record["decimals"] == 6
        assert "rounded to the stated decimals" in record["method"]
    # Two genuinely different models must not share a fingerprint.
    assert candidate["digest"] != champion["digest"]


def test_the_management_workbook_exists_and_states_its_boundaries() -> None:
    """An Excel workbook is what gets attached to an email, and it leaves the repository.

    So it has to carry the same boundaries the dashboard does rather than becoming the one
    surface where a number travels without its caveat.
    """
    from openpyxl import load_workbook

    workbook = load_workbook("docs/samples/monthly-fraud-kpi.xlsx")
    assert workbook.sheetnames == [
        "Summary",
        "Proposed approach",
        "Vendor performance",
        "Channel",
        "Boundaries",
    ]
    boundaries = " ".join(
        str(cell.value) for row in workbook["Boundaries"].iter_rows() for cell in row if cell.value
    )
    assert "no robust recommendation" in boundaries
    assert "not observed originations" in boundaries or "synthetic" in boundaries
    assert "not a verified vendor product" in boundaries

    proposed = " ".join(
        str(cell.value) for row in workbook["Proposed approach"].iter_rows() for cell in row if cell.value
    )
    assert "NOT approved" in proposed, "the rejected model must not read as a recommendation"

    # No placeholder values: every channel row carries a real reviewed count. Sheets carry
    # explanatory notes below the data, so rows are selected by their period label rather
    # than by being non-empty.
    channel = workbook["Channel"]
    reviewed = [
        row[5]
        for row in channel.iter_rows(min_row=4, max_col=6, values_only=True)
        if isinstance(row[0], str) and row[0].startswith("Period ")
    ]
    pack_reviewed = sum(
        row["good_customers_reviewed"]
        for row in load("evaluation/monthly_kpi.json")["channel"]
        if row["model_version"].startswith("incumbent")
    )
    assert reviewed and all(value is not None for value in reviewed)
    assert sum(reviewed) == pack_reviewed


def test_the_workbook_does_not_drift_from_the_database() -> None:
    from openpyxl import load_workbook

    pack = load("evaluation/monthly_kpi.json")
    incumbent = [row for row in pack["monthly"] if row["model_version"].startswith("incumbent")]
    summary = load_workbook("docs/samples/monthly-fraud-kpi.xlsx")["Summary"]
    rows = [
        row
        for row in summary.iter_rows(min_row=4, max_col=9, values_only=True)
        if isinstance(row[0], str) and row[0].startswith("Period ")
    ]
    assert len(rows) == len(incumbent)
    for sheet_row, record in zip(rows, incumbent, strict=True):
        assert sheet_row[0] == record["period"]
        assert sheet_row[1] == record["applications"]
        assert sheet_row[5] == record["fraud_caught"]
        assert abs(sheet_row[6] - record["catch_rate"]) < 1e-9


def test_the_sas_translations_claim_no_proficiency_and_carry_the_boundaries() -> None:
    readme = " ".join(Path("sas/README.md").read_text(encoding="utf-8").split())
    assert "translations, not production experience" in readme
    assert "nothing here has been run" in readme
    queue = Path("sas/review_queue.sas").read_text(encoding="utf-8")
    assert "none of which declines an applicant" in queue
    assert "no group attribute enters the ranking or the cut" in queue
    assert "Not executed" in queue
    assert "Not executed" in Path("sas/score_to_probability.sas").read_text(encoding="utf-8")


def test_the_defect_record_covers_every_investigation_with_its_regression_test() -> None:
    text = Path("docs/defect-record.md").read_text(encoding="utf-8")
    for defect in (
        "chance-level results",
        "could not score a single application",
        "measured zeros",
        "blind to the cheapest evasion",
    ):
        assert defect in text
    assert text.count("**Regression test.**") == 4
    assert text.count("**Root cause.**") >= 3
