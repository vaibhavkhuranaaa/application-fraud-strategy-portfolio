"""Monthly fraud KPI workbook and the management readout that goes with it.

The posting names Excel and PowerPoint as required tools and asks for regular fraud
performance reporting to management and business partners. The project produced CSVs and a
web page; neither is what gets attached to an email or opened in a meeting.

This builds the workbook. The readout is written as markdown rather than as a `.pptx`,
because `python-pptx` pulls in lxml and Pillow to produce a binary nobody can review in a
diff, and the readout's value is its content and its boundaries rather than its file
format. That is a decision, not an oversight, and it is recorded in `docs/management-readout.md`
and in the case study.

Every figure is read from `evaluation/monthly_kpi.json`, which is itself aggregated in
PostgreSQL from the fraud schema. Nothing is recomputed here, so the workbook cannot drift
from the database.

    PYTHONPATH=src uv run python scripts/build_management_pack.py
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from fraud_strategy.config import DEFAULT_EVIDENCE_DIR  # noqa: E402

OUTPUT_DIR = Path("docs/samples")
HEADER_FILL = PatternFill("solid", fgColor="1F2933")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=13)
NOTE_FONT = Font(italic=True, size=9, color="52606D")


def write_sheet(
    workbook: Workbook,
    title: str,
    heading: str,
    columns: list[str],
    rows: list[list[Any]],
    notes: list[str],
    formats: dict[int, str] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet["A1"] = heading
    sheet["A1"].font = TITLE_FONT
    header_row = 3
    for index, name in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for offset, values in enumerate(rows, start=header_row + 1):
        for index, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=index, value=value)
            if formats and index in formats:
                cell.number_format = formats[index]
    note_row = header_row + len(rows) + 2
    for offset, note in enumerate(notes):
        cell = sheet.cell(row=note_row + offset, column=1, value=note)
        cell.font = NOTE_FONT
    for index, name in enumerate(columns, start=1):
        width = max(len(name) + 2, 14)
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)


def build(evidence_dir: Path, output_dir: Path) -> dict[str, Any]:
    pack = json.loads((evidence_dir / "monthly_kpi.json").read_text(encoding="utf-8"))
    challenger = [row for row in pack["monthly"] if not row["model_version"].startswith("incumbent")]
    incumbent = [row for row in pack["monthly"] if row["model_version"].startswith("incumbent")]

    workbook = Workbook()
    workbook.remove(workbook.active)

    write_sheet(
        workbook,
        "Summary",
        "Monthly fraud KPI, incumbent score proxy",
        [
            "Period",
            "Applications",
            "Fraud attempts",
            "Attempt rate (bps)",
            "Cases worked",
            "Fraud caught",
            "Catch rate",
            "Reviewer hit rate",
            "Over capacity",
        ],
        [
            [
                r["period"],
                r["applications"],
                r["fraud_attempts"],
                r["fraud_rate_bps"],
                r["queue_size"],
                r["fraud_caught"],
                r["catch_rate"],
                r["investigator_yield"],
                r["capacity_overshoot"],
            ]
            for r in incumbent
        ],
        [
            "Reviewer hit rate is a leading indicator, available within days: a review confirms fraud "
            "at review time.",
            "Catch rate lags by 30 to 90 days and longer. It shares that numerator but divides by all "
            "fraud in the period, including what slipped past review. Recent periods will get worse as "
            "labels arrive.",
            "Over capacity is the queue above the staffed headcount. A score with ties cannot cut on an "
            "exact count, so the block at the cutting value cannot be split.",
            "Periods are relative source months, not calendar months.",
        ],
        formats={4: "0.0", 7: "0.0%", 8: "0.0%"},
    )

    write_sheet(
        workbook,
        "Proposed approach",
        "Same periods, scored by the proposed approach. Not approved for use.",
        [
            "Period",
            "Applications",
            "Fraud attempts",
            "Cases worked",
            "Fraud caught",
            "Catch rate",
            "Change vs prior",
            "Reviewer hit rate",
        ],
        [
            [
                r["period"],
                r["applications"],
                r["fraud_attempts"],
                r["queue_size"],
                r["fraud_caught"],
                r["catch_rate"],
                r["catch_rate_change"],
                r["investigator_yield"],
            ]
            for r in challenger
        ],
        [
            "This approach was evaluated and NOT approved. It fails two of eleven pre-agreed checks: "
            "calibration intercept and population stability.",
            "Shown so the size of the forgone benefit is visible, not as a recommendation.",
        ],
        formats={6: "0.0%", 7: "+0.0%;-0.0%;0.0%", 8: "0.0%"},
    )

    write_sheet(
        workbook,
        "Vendor performance",
        "Incumbent score proxy against the proposed approach, at identical review capacity",
        [
            "Period",
            "Incumbent catch rate",
            "Proposed catch rate",
            "Gap",
            "Incumbent caught",
            "Proposed caught",
            "Additional caught",
            "Cases worked",
        ],
        [
            [
                r["period"],
                r["incumbent_catch_rate"],
                r["challenger_catch_rate"],
                r["catch_rate_gap"],
                r["incumbent_fraud_caught"],
                r["challenger_fraud_caught"],
                r["additional_fraud_caught"],
                r["queue_size"],
            ]
            for r in pack["vendor_performance"]
        ],
        [
            "credit_risk_score is an incumbent score proxy standing in for a third-party decision "
            "score. It is not a verified vendor product and this is not a vendor service level.",
            "Both columns are measured at the same review capacity, so no part of the gap is bought "
            "with extra reviewers.",
        ],
        formats={2: "0.0%", 3: "0.0%", 4: "+0.0%;-0.0%;0.0%"},
    )

    write_sheet(
        workbook,
        "Channel",
        "Fraud by application channel, incumbent score proxy",
        ["Period", "Channel", "Applications", "Fraud attempts", "Fraud caught", "Good customers reviewed"],
        [
            [
                r["period"],
                r["channel"],
                r["applications"],
                r["fraud_attempts"],
                r["fraud_caught"],
                r["good_customers_reviewed"],
            ]
            for r in pack["channel"]
            if r["model_version"].startswith("incumbent")
        ],
        ["Groups with fewer than 200 fraud cases are not published anywhere in this program."],
    )

    write_sheet(
        workbook,
        "Boundaries",
        "What this pack is and is not",
        ["Statement"],
        [[line] for line in pack["limitations"]],
        [
            "Source: PostgreSQL fraud schema, analytics.fact_daily_strategy, via scripts/build_kpi_pack.py.",
            f"Dataset version: {pack['dataset_version']}.",
            "No period in this pack authorises a policy. The recorded strategy result is "
            "'no robust recommendation'.",
        ],
    )
    workbook["Boundaries"].column_dimensions["A"].width = 120
    for row in workbook["Boundaries"].iter_rows(min_row=4, max_col=1):
        row[0].alignment = Alignment(wrap_text=True, vertical="top")

    output_dir.mkdir(parents=True, exist_ok=True)
    destination = output_dir / "monthly-fraud-kpi.xlsx"
    workbook.save(destination)
    return {
        "workbook": str(destination),
        "sheets": workbook.sheetnames,
        "periods": len(challenger),
        "dataset_version": pack["dataset_version"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--evidence-dir", type=Path, default=DEFAULT_EVIDENCE_DIR)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    arguments = parser.parse_args()
    print(json.dumps(build(arguments.evidence_dir, arguments.output_dir), indent=2))


if __name__ == "__main__":
    main()
